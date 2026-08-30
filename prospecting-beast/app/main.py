import asyncio, json, os, time, uuid, secrets, io, csv
from pathlib import Path
from fastapi import FastAPI, HTTPException, Depends
from fastapi.responses import HTMLResponse, StreamingResponse, JSONResponse
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field, field_validator
from openpyxl import Workbook

from .core import *
from .osint_tools import xray_matrix, email_candidates_and_mx, fetch_public_page, parse_public_people

app = FastAPI(title='Prospecting Beast // Web Edition', version='32.0')
security = HTTPBasic(auto_error=False)
BASE = Path(__file__).resolve().parents[1]
app.mount('/static', StaticFiles(directory=BASE/'static'), name='static')

TASKS: dict[str, asyncio.Task] = {}


def auth(credentials: HTTPBasicCredentials | None = Depends(security)):
    pwd = env('APP_PASSWORD')
    if not pwd:
        raise HTTPException(500, 'APP_PASSWORD is not configured on the server.')
    if not credentials or not secrets.compare_digest(credentials.password, pwd):
        raise HTTPException(401, 'Authentication required', headers={'WWW-Authenticate': 'Basic'})
    return True


class JobIn(BaseModel):
    websites: list[str] = Field(min_length=1)
    mode: str = 'f'
    max_people_per_company: int = Field(default=50, ge=1, le=100)
    max_enrich_per_family: int = Field(default=20, ge=0, le=1000)
    max_related_per_company: int = Field(default=10, ge=0, le=50)
    max_companies_per_family: int = Field(default=100, ge=1, le=500)
    dry_run: bool = True
    excluded_people: list[str] = Field(default_factory=list)
    web_min_qualified: int = Field(default=5, ge=0, le=50)
    web_search_budget: int = Field(default=6, ge=1, le=20)
    location: str = Field(default='')
    industry: str = Field(default='')
    tavily_job_budget: int = Field(default=40, ge=1, le=500)
    osint_provider: str = 'web'
    spiderfoot_enabled: bool = False
    amass_enabled: bool = False
    agentic_discovery: bool = False

    @field_validator('osint_provider')
    @classmethod
    def validate_osint_provider(cls, v):
        v=str(v or 'web').lower().strip()
        if v not in ('web','web+osint'):
            v='web'
        return v

    @field_validator('mode')
    @classmethod
    def validate_mode(cls, v):
        v = v.lower().strip()
        if v not in ('f','nf','both'):
            raise ValueError('mode must be f, nf, or both')
        return v

    @field_validator('websites')
    @classmethod
    def validate_sites(cls, v):
        sites = unique_domains(v)
        if not sites:
            raise ValueError('At least one valid company domain is required.')
        max_roots = int(env('MAX_ROOT_COMPANIES','100'))
        if len(sites) > max_roots:
            raise ValueError(f'Maximum {max_roots} root companies per job.')
        return sites


def normalize_person_key(value: str) -> str:
    s = (value or '').strip().lower()
    s = ' '.join(s.split())
    return s


def parse_excluded_people(rows: list[str]) -> list[dict]:
    out = []
    for raw in rows or []:
        line = (raw or '').strip()
        if not line:
            continue
        parts = [x.strip() for x in line.split('|')]
        # Supported: Name | Email | LinkedIn OR Name | Domain | Email | LinkedIn
        item = {'name':'','domain':'','email':'','linkedin':''}
        if len(parts) == 1:
            item['name'] = parts[0]
        elif len(parts) == 2:
            item['name'], item['email'] = parts
        elif len(parts) == 3:
            item['name'], item['email'], item['linkedin'] = parts
        else:
            item['name'], item['domain'], item['email'], item['linkedin'] = parts[:4]
        for k in item:
            item[k] = normalize_person_key(item[k])
        if any(item.values()):
            out.append(item)
    # stable dedupe
    seen=set(); ded=[]
    for x in out:
        key=(x['name'],x['domain'],x['email'],x['linkedin'])
        if key not in seen:
            seen.add(key); ded.append(x)
    return ded


def person_is_excluded(person: dict, excluded: list[dict]) -> dict | None:
    name = normalize_person_key(person.get('name'))
    email = normalize_person_key(person.get('email'))
    linkedin = normalize_person_key(person.get('linkedin_url'))
    domain = normalize_person_key(person.get('domain'))
    for item in excluded or []:
        if item['linkedin'] and linkedin and item['linkedin'] == linkedin:
            return item
        if item['email'] and email and item['email'] == email:
            return item
        if item['name'] and name and item['name'] == name:
            if not item['domain'] or item['domain'] == domain:
                return item
    return None


async def load_lead_cache(job_id: str) -> dict[tuple[str,str,str,str], dict]:
    rows = await sb_select('leads', params={'job_id':f'eq.{job_id}','select':'linkedin,email,name,domain,phone,email,title,matched_title,company,raw,source,seamless_key_index','limit':'10000'})
    cache={}
    for r in rows or []:
        for key in lead_cache_keys(r):
            cache[key]=r
    return cache


def lead_cache_keys(p: dict) -> list[tuple[str,str,str,str]]:
    linkedin=normalize_person_key(p.get('linkedin_url'))
    email=normalize_person_key(p.get('email'))
    name=normalize_person_key(p.get('name'))
    domain=normalize_person_key(p.get('domain'))
    keys=[]
    if linkedin: keys.append((linkedin,'','',''))
    if email: keys.append(('',email,'',''))
    if name and domain: keys.append(('', '', name, domain))
    return keys


def rank_people(people: list[dict], mode: str, max_people: int) -> list[tuple[float,dict,dict]]:
    ranked=[]
    for p in people or []:
        title=p.get('title') or p.get('job_title') or ''
        info=base_score(title, mode)
        if not info:
            continue
        ranked.append((float(info.get('score',0)), info, p))
    ranked.sort(key=lambda x:(x[0], x[1].get('similarity',0), x[1].get('base_weight',0)), reverse=True)
    return ranked[:max_people]


async def get_job(jid: str):
    rows = await sb_select('jobs', params={'id': f'eq.{jid}', 'select':'*', 'limit':'1'})
    return rows[0] if rows else None


async def update_job(jid: str, **kwargs):
    kwargs['updated_at'] = 'now()'  # converted below
    patch = {}
    for k,v in kwargs.items():
        if k == 'updated_at' and v == 'now()':
            patch[k] = None
        elif k in ('stats','config','log') and not isinstance(v, str):
            patch[k] = v
        else:
            patch[k] = v
    # PostgREST cannot execute SQL expressions in JSON. Use actual timestamp.
    patch['updated_at'] = __import__('datetime').datetime.now(__import__('datetime').timezone.utc).isoformat()
    await sb_update('jobs', {'id':jid}, patch)


async def append_log(jid: str, message: str, level: str='info'):
    job = await get_job(jid)
    if not job:
        return
    logs = job.get('log') or []
    logs.append({'t': time.time(), 'level': level, 'm': message})
    logs = logs[-500:]
    await update_job(jid, log=logs)
    await sb_insert('logs', {'job_id':jid,'level':level,'message':message})


async def update_stats(jid: str, **updates):
    job = await get_job(jid)
    current = job.get('stats') or {}
    current.update(updates)
    await update_job(jid, stats=current)


def enrichment_meta(lead: dict) -> dict:
    raw = lead.get('raw') or {}
    return raw if isinstance(raw, dict) else {}


def enrichment_status(lead: dict) -> str:
    meta = enrichment_meta(lead)
    status = str(meta.get('_enrichment_status') or '').lower().strip()
    if status:
        return status
    if lead.get('seamless_key_index') or lead.get('phone') or lead.get('email'):
        return 'enriched'
    if meta.get('_enrichment_skipped') or str(lead.get('source') or '').endswith('skipped_enrichment'):
        return 'skipped'
    return 'ready'


def merge_enrichment_meta(lead: dict, **updates) -> dict:
    raw = enrichment_meta(lead).copy()
    raw.update(updates)
    return raw


async def load_job_leads(jid: str):
    return await sb_select('leads', params={
        'job_id':f'eq.{jid}',
        'select':'*',
        'order':'title_score.desc',
        'limit':'5000'
    })


async def family_enrichment_usage(jid: str) -> dict:
    rows = await sb_select('leads', params={'job_id':f'eq.{jid}','select':'family_id,seamless_key_index,source,raw','limit':'10000'})
    usage = {}
    for r in rows or []:
        if r.get('seamless_key_index'):
            fid = r.get('family_id') or ''
            usage[fid] = usage.get(fid, 0) + 1
    return usage


ENRICH_TASKS: dict[str, asyncio.Task] = {}


async def run_manual_enrichment(job_id: str, lead_ids: list[str]):
    pool = SeamlessPool()
    job = await get_job(job_id)
    if not job:
        raise RuntimeError('Job not found')
    cfg = job.get('config') or {}
    family_budget = int(cfg.get('max_enrich_per_family',20) or 0)
    excluded = cfg.get('excluded_people') or []
    if excluded and isinstance(excluded[0], str):
        excluded = parse_excluded_people(excluded)
    rows = await load_job_leads(job_id)
    wanted = {str(x) for x in lead_ids}
    selected_rows = [r for r in rows if str(r.get('id')) in wanted]
    if not selected_rows:
        return {'processed':0,'skipped':0,'failed':0,'message':'No selected leads found.'}
    usage = await family_enrichment_usage(job_id)
    processed=skipped=failed=0
    await append_log(job_id, f'Manual Seamless enrichment started for {len(selected_rows)} selected leads.', 'info')
    for lead in selected_rows:
        if await job_stop_requested(job_id):
            await append_log(job_id, 'Manual Seamless enrichment stopped by user.', 'warning')
            break
        status=enrichment_status(lead)
        if status in {'enriched','skipped','running','queued'}:
            skipped += 1
            continue
        # Respect persistent skip list even if the lead came from a prior run.
        person = {
            'name':lead.get('name',''),
            'title':lead.get('title',''),
            'domain':lead.get('domain',''),
            'company':lead.get('company',''),
            'linkedin_url':lead.get('linkedin',''),
            'email':lead.get('email',''),
            'phone':lead.get('phone','')
        }
        if person_is_excluded(person, excluded):
            await sb_update('leads', {'id':lead['id']}, {'raw':merge_enrichment_meta(lead, _enrichment_status='skipped', _enrichment_skipped=True, _skip_reason='Already have this contact', _enrichment_updated_at=time.time()), 'source':'public_web+skipped_enrichment'})
            skipped += 1
            continue
        fid = lead.get('family_id') or ''
        if usage.get(fid,0) >= family_budget:
            await sb_update('leads', {'id':lead['id']}, {'raw':merge_enrichment_meta(lead, _enrichment_status='quota_blocked', _enrichment_error=f'Family Seamless budget {family_budget} reached', _enrichment_updated_at=time.time())})
            await append_log(job_id, f'SKIP Seamless — {lead.get("name")}: family budget reached ({usage.get(fid,0)}/{family_budget}).', 'warning')
            skipped += 1
            continue
        await sb_update('leads', {'id':lead['id']}, {'raw':merge_enrichment_meta(lead, _enrichment_status='running', _enrichment_updated_at=time.time())})
        await append_log(job_id, f'Seamless: enriching {lead.get("name")} — {lead.get("title")}.', 'info')
        try:
            result = await pool.research(person)
            phone = result.get('contactPhone1') or result.get('phone') or result.get('contactPhone2') or result.get('directDial') or ''
            email = result.get('email') or result.get('email1') or result.get('workEmail') or result.get('personalEmail') or ''
            key_idx = result.get('_seamless_key_index')
            status = 'enriched' if (phone or email) else 'no_data'
            raw = merge_enrichment_meta(lead, _enrichment_status=status, _enrichment_updated_at=time.time(), _seamless_key_index=key_idx)
            await sb_update('leads', {'id':lead['id']}, {'phone':phone,'email':email,'source':'public_web+seamless','seamless_key_index':key_idx,'raw':raw})
            usage[fid]=usage.get(fid,0)+1
            processed += 1
            await update_stats(job_id, seamless_credits_used=sum(usage.values()), seamless_keys_active=len(pool.available_indices()))
            await append_log(job_id, f'Seamless OK — {lead.get("name")}: phone={bool(phone)} email={bool(email)} key={key_idx or "?"}.', 'success')
        except Exception as e:
            failed += 1
            await sb_update('leads', {'id':lead['id']}, {'raw':merge_enrichment_meta(lead, _enrichment_status='failed', _enrichment_error=str(e), _enrichment_updated_at=time.time())})
            await append_log(job_id, f'SEAMLESS ERROR — {lead.get("name")}: {e}', 'error')
    await append_log(job_id, f'Manual Seamless enrichment complete: {processed} enriched, {skipped} skipped, {failed} failed.', 'info')
    return {'processed':processed,'skipped':skipped,'failed':failed}


async def find_or_create_company(job_id, family_id, root_domain, domain, relationship='original', parent_domain=None, name=None):
    rows = await sb_select('companies', params={'job_id':f'eq.{job_id}','domain':f'eq.{domain}','select':'*','limit':'1'})
    if rows:
        return rows[0]
    return await sb_insert('companies', {
        'job_id':job_id,'family_id':family_id,'root_domain':root_domain,'domain':domain,
        'name':name or domain,'relationship':relationship,'parent_domain':parent_domain,
        'status':'queued','evidence':[]
    })


async def save_relationship(job_id, family_id, source_domain, rr):
    rd = norm_domain(rr.get('domain',''))
    if not rd:
        return
    try:
        await sb_insert('relationships', {
            'job_id':job_id,'family_id':family_id,'source_domain':source_domain,
            'related_domain':rd,'related_name':rr.get('company') or rr.get('name') or rd,
            'relationship':rr.get('relationship','related'),'confidence':rr.get('confidence',0),
            'evidence':rr.get('evidence',[]) or []
        })
    except Exception:
        pass


async def save_lead(job_id, family_id, company_id, p, lead_cache: dict[tuple[str,str,str,str],dict] | None = None):
    cache = lead_cache if lead_cache is not None else {}
    # Application-side dedupe. This avoids fragile PostgREST filter strings and is faster within a job.
    if any(k in cache for k in lead_cache_keys(p)):
        return False
    domain=(p.get('domain') or '').strip().lower()
    row={
        'job_id':job_id,'family_id':family_id,'company_id':company_id,
        'company':p.get('company',''),'domain':domain,'root_domain':p.get('root_domain',''),'parent_domain':p.get('parent_domain',''),'relationship':p.get('relationship',''),'relationship_display':p.get('relationship_display',''),
        'name':p.get('name',''),'title':p.get('title',''),'matched_title':p.get('matched_title',''),
        'title_score':p.get('score',0),'similarity':p.get('similarity',0),
        'linkedin':p.get('linkedin_url',''),'phone':p.get('phone',''),'email':p.get('email',''),
        'source':p.get('source','public_web'),'seamless_key_index':p.get('_seamless_key_index'),
        'raw':{**(p.get('raw',{}) or {}), '_match_type':p.get('match_type','Related'), '_match_reason':p.get('match_reason',''), '_relationship_display':p.get('relationship_display','')}
    }
    await sb_insert('leads', row)
    for k in lead_cache_keys(p):
        cache[k]=row
    return True


def relationship_display(relationship: str, source_domain: str) -> str:
    r=(relationship or 'original').lower().strip()
    if r=='original': return 'Original company'
    return {'parent':f'Parent of {source_domain}','subsidiary':f'Subsidiary of {source_domain}','sister':f'Sister company of {source_domain}','brand':f'Brand / business unit of {source_domain}','acquired_target':f'Acquired / formerly separate from {source_domain}','related':f'Related company to {source_domain}'}.get(r,r.title())

def extract_person(p, domain, relationship, scoreinfo, parent_domain=None, root_domain=None):
    org = p.get('organization')
    company = org.get('name') if isinstance(org,dict) else p.get('organization_name') or p.get('company_name') or domain
    name = p.get('name') or p.get('full_name') or ' '.join(x for x in [p.get('first_name'),p.get('last_name')] if x).strip()
    lineage_target = parent_domain or root_domain or domain
    return {
        'name':name,
        'title':p.get('title') or p.get('job_title') or '',
        'linkedin_url':p.get('linkedin_url') or p.get('linkedin_url_normalized') or '',
        'domain':domain,'company':company,'relationship':relationship,'relationship_display':relationship_display(relationship,lineage_target),
        'score':scoreinfo['score'],'matched_title':scoreinfo['matched_title'],'similarity':scoreinfo['similarity'],
        'match_type':scoreinfo.get('match_type','Related'),'match_reason':scoreinfo.get('match_reason',''),
        'source':'public_web','raw':p
    }


async def job_stop_requested(jid: str) -> bool:
    job = await get_job(jid)
    if not job:
        return True
    status = (job.get('status') or '').lower()
    cfg = job.get('config') or {}
    return status in ('stopping','stopped','cancelled') or bool(cfg.get('stop_requested'))


async def mark_job_stopped(jid: str, reason: str = 'Stopped by user.'):
    job = await get_job(jid)
    if not job:
        return
    cfg = job.get('config') or {}
    cfg['stop_requested'] = True
    await update_job(jid, status='stopped', config=cfg, error=None)
    await append_log(jid, reason, 'warning')


async def process_company(job_id, family_id, root_domain, domain, relationship, parent_domain, cfg, seamless_pool, family_state, lead_cache):
    company = await find_or_create_company(job_id,family_id,root_domain,domain,relationship,parent_domain)
    company_id = company['id']
    now = __import__('datetime').datetime.now(__import__('datetime').timezone.utc).isoformat()
    await sb_update('companies', {'id':company_id}, {'status':'processing','updated_at':now})
    await append_log(job_id, f'[{domain}] START — relationship={relationship}', 'info')
    if await job_stop_requested(job_id):
        await sb_update('companies', {'id':company_id}, {'status':'stopped','updated_at':now})
        return []

    company_name = company.get('name') or domain
    await append_log(job_id, f'[{domain}] WEB: public people discovery…', 'info')
    job_before = await get_job(job_id)
    stats_before = (job_before or {}).get('stats') or {}
    used_tavily = int(stats_before.get('tavily_estimated_credits', 0) or 0)
    tavily_budget = int(cfg.get('tavily_job_budget', 40) or 40)
    remaining_tavily = max(0, tavily_budget - used_tavily)
    if remaining_tavily <= 0:
        await append_log(job_id, f'[{domain}] Tavily budget exhausted; skipping further web search.', 'warning')
        web = {'people': [], 'estimated_tavily_credits': 0, 'evidence_count': 0, 'query_count': 0, 'target_terms': list(title_dictionary(cfg['mode']).keys()), 'sources_used': []}
    else:
        max_queries = min(int(cfg.get('web_search_budget', 6) or 6), remaining_tavily)
        web = await web_people_discovery(company_name, domain, cfg['mode'], min(cfg['max_people_per_company'], 25), max_queries, cfg.get('location',''), cfg.get('industry',''), bool(cfg.get('agentic_discovery',False)))
    people = web.get('people') or []
    tavily_spent = int(web.get('estimated_tavily_credits', 0) or 0)
    if tavily_spent:
        job_now = await get_job(job_id)
        st_now = (job_now or {}).get('stats') or {}
        st_now['tavily_estimated_credits'] = int(st_now.get('tavily_estimated_credits', 0) or 0) + tavily_spent
        await update_job(job_id, stats=st_now)
        if int(st_now.get('tavily_estimated_credits', 0)) >= int(cfg.get('tavily_job_budget', 40)):
            await append_log(job_id, f'[{domain}] Tavily job budget reached ({cfg.get("tavily_job_budget",40)} estimated credits).', 'warning')
    qcount = sum(1 for p in people if base_score(p.get('title',''), cfg['mode']))
    await append_log(job_id, f'[{domain}] WEB OK — {len(people)} candidates; {qcount} pass title matching.', 'success' if people else 'warning')
    normalized=[]
    for p in people:
        if p.get('job_title') or p.get('full_name'):
            normalized.append(p)
        else:
            normalized.append({'name':p.get('name'),'title':p.get('title'),'job_title':p.get('title'),'linkedin_url':p.get('linkedin_url',''),'organization_name':company_name,'job_company_name':company_name,'_source':'public_web','source_url':p.get('_source_url','')})
    people=normalized
    ranked = rank_people(people, cfg['mode'], min(cfg['max_people_per_company'], len(people) or 0))
    now_iso=__import__('datetime').datetime.now(__import__('datetime').timezone.utc).isoformat()
    await sb_update('companies', {'id':company_id}, {'people_found':len(people),'qualified_people':len(ranked),'updated_at':now_iso})

    # IMPORTANT: discovery never calls Seamless automatically. Qualified leads are
    # stored as ready/skipped and enriched only when the user explicitly clicks
    # “Seamless Enrich”.
    saved_here=0
    for _, info, raw_person in ranked:
        if await job_stop_requested(job_id):
            await append_log(job_id, f'[{domain}] Stop requested — skipping remaining lead saves.', 'warning')
            break
        person = extract_person(raw_person,domain,relationship,info,parent_domain=parent_domain,root_domain=root_domain)
        person['root_domain']=root_domain
        person['parent_domain']=parent_domain or ''
        if not person['name']:
            continue
        cached_row = next((lead_cache[k] for k in lead_cache_keys(person) if k in lead_cache), None)
        if cached_row:
            await append_log(job_id, f'[{domain}] CACHE HIT — {person["name"]}; existing lead kept, no Seamless call.', 'info')
            continue
        excluded = person_is_excluded(person, cfg.get('excluded_people', []))
        meta = {'_enrichment_status':'skipped' if excluded else 'ready', '_enrichment_updated_at':now_iso}
        if excluded:
            meta['_enrichment_skipped'] = True
            meta['_skip_reason'] = 'Already have this contact'
        person['phone']=''; person['email']=''
        person['source']='public_web+skipped_enrichment' if excluded else 'public_web'
        person['raw'] = {**(person.get('raw',{}) or {}), **meta}
        if await save_lead(job_id,family_id,company_id,person,lead_cache):
            saved_here += 1
            family_state['leads_saved'] = family_state.get('leads_saved',0) + 1
            if excluded:
                await append_log(job_id, f'[{domain}] SKIP Seamless — {person["name"]} is on the no-enrichment list.', 'info')

    await sb_update('companies', {'id':company_id}, {'enriched_people':0,'updated_at':now_iso})

    if await job_stop_requested(job_id):
        await sb_update('companies', {'id':company_id}, {'status':'stopped','updated_at':__import__('datetime').datetime.now(__import__('datetime').timezone.utc).isoformat()})
        return []

    # Corporate family discovery never uses Seamless credits.
    company_name = next((extract_person(r,domain,relationship,i,parent_domain=parent_domain,root_domain=root_domain)['company'] for _,i,r in ranked if r), domain)
    job_now = await get_job(job_id)
    st_now = (job_now or {}).get('stats') or {}
    if int(st_now.get('tavily_estimated_credits', 0) or 0) >= int(cfg.get('tavily_job_budget', 40)):
        await append_log(job_id, f'[{domain}] Tavily job budget exhausted; skipping further corporate-family web research.', 'warning')
        related = []
        await sb_update('companies', {'id':company_id}, {'status':'done' if ranked or people else 'done_no_people','evidence':[],'updated_at':__import__('datetime').datetime.now(__import__('datetime').timezone.utc).isoformat()})
        return related
    await append_log(job_id, f'[{domain}] Corporate-family research: searching web evidence…', 'info')
    try:
        result = await discover_family(company_name, domain)
        # v27+ contract: always returns (relationships, estimated_tavily_credits).
        if isinstance(result, tuple) and len(result) == 2:
            related, family_tavily = result
        elif isinstance(result, list):
            related, family_tavily = result, 0
        else:
            raise RuntimeError(f'Unexpected family discovery result: {type(result).__name__}')
        job_now = await get_job(job_id)
        st_now = (job_now or {}).get('stats') or {}
        st_now['tavily_estimated_credits'] = int(st_now.get('tavily_estimated_credits', 0) or 0) + int(family_tavily or 0)
        await update_job(job_id, stats=st_now)
        await append_log(job_id, f'[{domain}] Corporate-family research OK — {len(related)} relationship candidates. Tavily est. +{family_tavily}.', 'success')
    except Exception as e:
        related = []
        await append_log(job_id, f'[{domain}] FAMILY RESEARCH ERROR: {e}', 'error')

    status='done' if ranked or people else 'done_no_people'
    await sb_update('companies', {'id':company_id}, {'status':status,'evidence':related,'updated_at':__import__('datetime').datetime.now(__import__('datetime').timezone.utc).isoformat()})
    return related


async def process_job(jid: str, cfg: dict):
    try:
        await update_job(jid,status='running')
        await append_log(jid,'Job started')
        pool = SeamlessPool()
        lead_cache = await load_lead_cache(jid)
        roots = unique_domains(cfg['websites'])
        existing = await sb_select('companies', params={'job_id':f'eq.{jid}','select':'*','order':'created_at.asc','limit':'5000'})
        by_root={}
        for c in existing or []:
            by_root.setdefault(c.get('root_domain') or '', []).append(c)
        cfg.setdefault('family_credit_usage', {})
        existing_job = await get_job(jid)
        existing_stats = (existing_job or {}).get('stats') or {}
        total_companies = int(existing_stats.get('companies_processed',0) or 0)
        total_credits = int(existing_stats.get('seamless_credits_used',0) or 0)
        total_leads = int(existing_stats.get('leads_saved',0) or 0)
        root_index = 0
        for root in roots:
            if await job_stop_requested(jid):
                await mark_job_stopped(jid, 'Stop requested — no new families will be started.')
                return
            root_index += 1
            rows = by_root.get(root, [])
            family_id = rows[0].get('family_id') if rows and rows[0].get('family_id') else str(uuid.uuid4())
            completed={c.get('domain') for c in rows if c.get('status') in ('done','done_no_people')}
            queue=[]
            if rows:
                for c in rows:
                    status=(c.get('status') or 'queued').lower()
                    if c.get('domain') and c.get('domain') not in completed and status in ('queued','processing','error','stopped'):
                        queue.append((c['domain'],c.get('relationship','related'),c.get('parent_domain')))
                if not queue and root not in completed:
                    queue=[(root,'original',None)]
            else:
                queue=[(root,'original',None)]
            family_state={'credits_used':int((cfg.get('family_credit_usage') or {}).get(root,0) or 0),'leads_saved':0}
            family_companies=0
            last_domain=root
            await append_log(jid,f'Starting family {root} ({root_index}/{len(roots)}) — resuming {len(queue)} queued companies.' if rows else f'Starting family {root} ({root_index}/{len(roots)})')
            while queue and family_companies < cfg['max_companies_per_family']:
                if await job_stop_requested(jid):
                    await mark_job_stopped(jid, f'Family {root} stopped by user.')
                    return
                domain, relationship, parent_domain = queue.pop(0)
                last_domain=domain
                if domain in completed:
                    continue
                family_companies += 1
                total_companies += 1
                await update_stats(jid, root_companies=len(roots), companies_processed=total_companies, current_company=domain, current_root=root, family_credits_used=family_state['credits_used'], leads_saved=total_leads)
                try:
                    before_company_credits = family_state['credits_used']
                    before_company_leads = family_state.get('leads_saved',0)
                    related = await process_company(jid,family_id,root,domain,relationship,parent_domain,cfg,pool,family_state,lead_cache)
                    total_credits += max(0, family_state['credits_used'] - before_company_credits)
                    total_leads += max(0, family_state.get('leads_saved',0) - before_company_leads)
                    cfg['family_credit_usage'][root]=family_state['credits_used']
                    await update_job(jid, config=cfg)
                except Exception as e:
                    await append_log(jid,f'{domain}: ERROR {e}','error')
                    row = await sb_select('companies',params={'job_id':f'eq.{jid}','domain':f'eq.{domain}','select':'id','limit':'1'})
                    if row:
                        await sb_update('companies',{'id':row[0]['id']},{'status':'error','updated_at':__import__('datetime').datetime.now(__import__('datetime').timezone.utc).isoformat()})
                    continue
                added=0
                for rr in related or []:
                    rd=norm_domain(rr.get('domain',''))
                    conf=float(rr.get('confidence',0) or 0)
                    rel=(rr.get('relationship') or 'related').lower().strip()
                    queue_score=float(rr.get('queue_score') or 0)
                    await save_relationship(jid,family_id,domain,rr)
                    # Only queue high-value corporate relationships for prospecting.
                    # Keep lower-value relationships in the map, but don't let brands /
                    # vague 'related' entities explode the family queue.
                    min_conf=float(env('MIN_RELATIONSHIP_CONFIDENCE','0.72'))
                    min_queue_score=float(env('MIN_RELATIONSHIP_QUEUE_SCORE','55'))
                    if not rd or rd==domain or rd in completed or any(x[0]==rd for x in queue) or conf < min_conf or queue_score < min_queue_score:
                        continue
                    if added >= cfg['max_related_per_company']:
                        continue
                    queue.append((rd,rel,domain))
                    await find_or_create_company(jid,family_id,root,rd,rel,domain,rr.get('company') or rr.get('name'))
                    added += 1
                await append_log(jid,f'{domain}: corporate research found {len(related or [])} candidates; queued {added}.')
            await update_stats(jid, root_companies=len(roots), companies_processed=total_companies, seamless_credits_used=total_credits, current_company=last_domain, current_root=root, leads_saved=total_leads)
            await append_log(jid,f'Family {root} complete. Seamless research used {family_state["credits_used"]}/{cfg["max_enrich_per_family"]}. Leads saved this family: {family_state.get("leads_saved",0)}.')
        job = await get_job(jid); st=job.get('stats') or {}; st.update({'root_companies':len(roots),'companies_processed':total_companies,'seamless_credits_used':total_credits,'leads_saved':total_leads})
        await update_job(jid,status='completed',stats=st)
        await append_log(jid,'Job completed')
    except Exception as e:
        await update_job(jid,status='failed',error=str(e))
        await append_log(jid,f'JOB ERROR: {e}','error')
    finally:
        TASKS.pop(jid,None)


async def resume_incomplete_jobs():
    try:
        rows = await sb_select('jobs', params={'status':'in.(queued,running)','select':'*','limit':'20'})
    except Exception:
        return
    for j in rows:
        if j['id'] in TASKS:
            continue
        try:
            cfg=j.get('config') or {}
            TASKS[j['id']] = asyncio.create_task(process_job(j['id'],cfg))
        except Exception:
            pass


@app.on_event('startup')
async def startup():
    await resume_incomplete_jobs()


@app.get('/', response_class=HTMLResponse)
async def home(_: bool = Depends(auth)):
    from fastapi.responses import Response
    html=(BASE/'web'/'index.html').read_text(encoding='utf-8')
    return Response(content=html, media_type='text/html', headers={'Cache-Control':'no-store, no-cache, must-revalidate, max-age=0'})


@app.get('/api/health')
async def health(_: bool = Depends(auth)):
    pool=SeamlessPool()
    return {'ok':True,'seamless_keys_configured':len(pool.keys),'seamless_keys_available':len(pool.available_indices()),'supabase':supabase_enabled(),'gemini_model':env('GEMINI_MODEL','gemini-3.5-flash-lite'),'tavily_configured':bool(env('TAVILY_API_KEY')),'discovery_mode':'web_only','spiderfoot_enabled':str(env('SPIDERFOOT_ENABLED','false')).lower() in {'1','true','yes','on'},'amass_enabled':str(env('AMASS_ENABLED','false')).lower() in {'1','true','yes','on'},'tavily_monthly_budget_hint':int(env('TAVILY_MONTHLY_BUDGET','1500')), 'agentic_discovery':str(env('AGENTIC_DISCOVERY','false')).lower() in {'1','true','yes','on'}}



@app.get('/api/jobs/{jid}/diagnostics')
async def job_diagnostics(jid: str, _: bool = Depends(auth)):
    job = await get_job(jid)
    if not job:
        raise HTTPException(404, 'Job not found')
    companies = await sb_select('companies', params={'job_id':f'eq.{jid}','select':'*','order':'created_at.asc'})
    logs = await sb_select('logs', params={'job_id':f'eq.{jid}','select':'*','order':'created_at.desc','limit':'200'})
    return {'job': job, 'companies': companies, 'logs': list(reversed(logs or []))}


@app.post('/api/test/web-discovery')
async def test_web_discovery(payload: dict, _: bool = Depends(auth)):
    domain = norm_domain(str(payload.get('domain','')))
    mode = str(payload.get('mode','f')).lower()
    max_people = max(1, min(int(payload.get('max_people',10)), 25))
    if not domain:
        raise HTTPException(400, 'domain is required')
    try:
        # Public-web discovery never calls Seamless.
        out = await web_people_discovery(domain, domain, mode, max_people, 8, str(payload.get('location','') or ''), str(payload.get('industry','') or ''), bool(payload.get('agentic',False)))
        people = out.get('people') or []
        return {
            'ok': True,
            'domain': domain,
            'count': len(people),
            'stats': {'people_returned': len(people), 'evidence_count': out.get('evidence_count',0)},
            'evidence_count': out.get('evidence_count', 0),
            'people': [
                {
                    'name': p.get('name',''),
                    'title': p.get('title',''),
                    'linkedin': p.get('linkedin_url',''),
                    'source_url': p.get('_source_url',''),
                    'score': base_score(p.get('title',''), mode),
                } for p in people
            ],
            'tavily_configured': bool(env('TAVILY_API_KEY')),
            'gemini_configured': bool(env('GEMINI_API_KEY')),
        }
    except Exception as e:
        return JSONResponse(status_code=200, content={'ok': False, 'domain': domain, 'error': f'{type(e).__name__}: {e}'})

@app.post('/api/jobs')
async def create_job(inp: JobIn, _: bool = Depends(auth)):
    try:
        cfg=inp.model_dump()
        cfg['excluded_people'] = parse_excluded_people(cfg.get('excluded_people', []))
        if cfg['dry_run'] is None:
            cfg['dry_run']=env('DRY_RUN','true').lower() == 'true'
        jid=str(uuid.uuid4())
        await sb_insert('jobs', {'id':jid,'status':'queued','mode':cfg['mode'],'config':cfg,'stats':{'root_companies':len(cfg['websites']),'companies_processed':0,'seamless_credits_used':0},'log':[]})
        TASKS[jid]=asyncio.create_task(process_job(jid,cfg))
        return {'id':jid}
    except Exception as e:
        return JSONResponse(status_code=200, content={'ok':False,'error':f'{type(e).__name__}: {e}'})


@app.get('/api/jobs')
async def list_jobs(include_archived: bool = False, hide_empty: bool = False, _: bool = Depends(auth)):
    # New schema has archived_at. Fall back to legacy schema if the migration has not
    # been applied yet so the dashboard stays usable while the user updates Supabase.
    try:
        params={'select':'id,status,mode,created_at,updated_at,stats,error,archived_at','order':'created_at.desc','limit':'100'}
        if not include_archived:
            params['archived_at']='is.null'
        rows=await sb_select('jobs', params=params)
    except Exception:
        params={'select':'id,status,mode,created_at,updated_at,stats,error','order':'created_at.desc','limit':'100'}
        rows=await sb_select('jobs', params=params)
    if hide_empty:
        out=[]
        for j in rows or []:
            st=(j.get('status') or '').lower()
            stats=j.get('stats') or {}
            leads=int(stats.get('leads_saved') or stats.get('leads_found') or stats.get('qualified_people') or 0)
            if leads==0 and st in {'completed','failed','stopped'}:
                continue
            out.append(j)
        rows=out
    return rows


@app.post('/api/jobs/{jid}/archive')
async def archive_job(jid: str, _: bool = Depends(auth)):
    job=await get_job(jid)
    if not job: raise HTTPException(404,'job not found')
    if (job.get('status') or '').lower() in ('queued','running','stopping'):
        raise HTTPException(409,'Cannot archive a running job. Stop it first.')
    try:
        await update_job(jid, archived_at=__import__('datetime').datetime.now(__import__('datetime').timezone.utc).isoformat())
    except Exception as e:
        raise HTTPException(500, f'Archive requires the latest jobs schema migration: {e}')
    await append_log(jid,'Job archived from the active queue.','info')
    return {'ok':True,'status':'archived'}


@app.post('/api/jobs/archive-empty')
async def archive_empty_jobs(_: bool = Depends(auth)):
    try:
        rows=await sb_select('jobs', params={'select':'id,status,stats,archived_at','archived_at':'is.null','limit':'500'})
    except Exception as e:
        raise HTTPException(500, f'Queue cleanup requires the latest jobs schema migration: {e}')
    archived=0
    for j in rows or []:
        st=(j.get('status') or '').lower(); stats=j.get('stats') or {}
        leads=int(stats.get('leads_saved') or stats.get('leads_found') or stats.get('qualified_people') or 0)
        if leads==0 and st in {'completed','failed','stopped'}:
            await sb_update('jobs', {'id':j['id']}, {'archived_at':__import__('datetime').datetime.now(__import__('datetime').timezone.utc).isoformat()})
            archived+=1
    return {'ok':True,'archived':archived}


@app.post('/api/jobs/{jid}/unarchive')
async def unarchive_job(jid: str, _: bool = Depends(auth)):
    job=await get_job(jid)
    if not job: raise HTTPException(404,'job not found')
    try:
        await sb_update('jobs', {'id':jid}, {'archived_at':None})
    except Exception as e:
        raise HTTPException(500, f'Unarchive requires the latest jobs schema migration: {e}')
    return {'ok':True}


@app.delete('/api/jobs/{jid}')
async def delete_job(jid: str, _: bool = Depends(auth)):
    job=await get_job(jid)
    if not job: raise HTTPException(404,'job not found')
    if jid in TASKS and not TASKS[jid].done():
        raise HTTPException(409,'Cannot delete a running job. Stop it first.')
    # Supabase helper added in core.py; delete cascades companies/leads/relationships/logs.
    try:
        await sb_delete('jobs', {'id':jid})
    except Exception as e:
        raise HTTPException(500, f'Delete failed: {e}')
    TASKS.pop(jid,None)
    return {'ok':True}


@app.get('/api/jobs/{jid}')
async def get_job_detail(jid: str, _: bool = Depends(auth)):
    try:
        job=await get_job(jid)
        if not job: raise HTTPException(404,'job not found')
        leads=await sb_select('leads', params={'job_id':f'eq.{jid}','select':'*','order':'title_score.desc','limit':'2000'})
        companies=await sb_select('companies', params={'job_id':f'eq.{jid}','select':'id,family_id,root_domain,domain,name,relationship,parent_domain,status,people_found,qualified_people,enriched_people,updated_at','order':'updated_at.desc','limit':'1000'})
        relationships=await sb_select('relationships', params={'job_id':f'eq.{jid}','select':'source_domain,related_domain,related_name,relationship,confidence,evidence','order':'confidence.desc','limit':'1000'})
        logs=await sb_select('logs', params={'job_id':f'eq.{jid}','select':'created_at,level,message','order':'created_at.desc','limit':'100'})
        return {'job':job,'leads':leads,'companies':companies,'relationships':relationships,'logs':list(reversed(logs))}
    except HTTPException:
        raise
    except Exception as e:
        return JSONResponse(status_code=200, content={'ok':False,'error':f'{type(e).__name__}: {e}'})


class EnrichIn(BaseModel):
    lead_ids: list[str] = Field(min_length=1, max_length=100)


@app.post('/api/jobs/{jid}/enrich')
async def enrich_selected(jid: str, inp: EnrichIn, _: bool = Depends(auth)):
    job=await get_job(jid)
    if not job: raise HTTPException(404,'job not found')
    existing = ENRICH_TASKS.get(jid)
    if existing and not existing.done():
        raise HTTPException(409,'Seamless enrichment is already running for this job.')
    # Do not require the discovery worker to be running. Manual enrichment is explicitly user-triggered.
    ENRICH_TASKS[jid]=asyncio.create_task(run_manual_enrichment(jid, inp.lead_ids))
    return {'ok':True,'queued':len(inp.lead_ids),'message':'Manual Seamless enrichment started.'}


@app.post('/api/jobs/{jid}/enrich/stop')
async def stop_enrichment(jid: str, _: bool = Depends(auth)):
    task=ENRICH_TASKS.get(jid)
    if task and not task.done():
        task.cancel()
        await append_log(jid,'Manual Seamless enrichment cancellation requested.','warning')
        return {'ok':True,'status':'stopping'}
    return {'ok':True,'status':'idle'}


@app.post('/api/jobs/{jid}/stop')
async def stop_job(jid: str, _: bool = Depends(auth)):
    job=await get_job(jid)
    if not job: raise HTTPException(404,'job not found')
    status=(job.get('status') or '').lower()
    if status not in ('queued','running'):
        return {'ok':True,'status':status}
    cfg=job.get('config') or {}
    cfg['stop_requested']=True
    await update_job(jid,status='stopping',config=cfg)
    await append_log(jid,'Stop requested. The worker will finish the current network call and stop before starting another company.', 'warning')
    return {'ok':True,'status':'stopping'}


@app.post('/api/jobs/{jid}/resume')
async def resume_job(jid: str, _: bool = Depends(auth)):
    job=await get_job(jid)
    if not job: raise HTTPException(404,'job not found')
    cfg=job.get('config') or {}
    cfg['stop_requested']=False
    # Resume only when no active worker exists.
    if jid not in TASKS or TASKS[jid].done():
        await update_job(jid,status='queued',config=cfg,error=None)
        TASKS[jid]=asyncio.create_task(process_job(jid,cfg))
    else:
        await update_job(jid,status='running',config=cfg,error=None)
    await append_log(jid,'Job resumed.')
    return {'ok':True}


class XRayToolIn(BaseModel):
    company: str = ''
    domain: str = ''
    mode: str = 'f'
    location: str = ''
    industry: str = ''
    limit: int = Field(default=40, ge=1, le=100)

@app.post('/api/tools/xray-matrix')
async def api_xray_matrix(inp: XRayToolIn, _: bool = Depends(auth)):
    domain=norm_domain(inp.domain or inp.company)
    company=(inp.company or domain).strip()
    roles=list(title_dictionary(inp.mode).keys())
    rows=xray_matrix(company,[domain] if domain else [],roles,inp.location,inp.industry)
    return {'ok':True,'queries':rows[:inp.limit],'count':min(len(rows),inp.limit)}

class EmailToolIn(BaseModel):
    first_name: str
    last_name: str
    domain: str

@app.post('/api/tools/email-candidates')
async def api_email_candidates(inp: EmailToolIn, _: bool = Depends(auth)):
    domain=norm_domain(inp.domain)
    if not domain: raise HTTPException(400,'valid company domain is required')
    return {'ok':True, **(await email_candidates_and_mx(inp.first_name, inp.last_name, domain))}

class PublicParseIn(BaseModel):
    urls: list[str] = Field(min_length=1, max_length=10)

@app.post('/api/tools/public-page-parse')
async def api_public_page_parse(inp: PublicParseIn, _: bool = Depends(auth)):
    rows=[]
    for raw in inp.urls:
        url=(raw or '').strip()
        if not url: continue
        try:
            page=await fetch_public_page(url)
            people=parse_public_people(page)
            rows.append({'url':url,'ok':True,'title':page.title,'people':people,'emails':extract_public_emails(page.text),'json_ld_count':len(page.json_ld)})
        except Exception as exc:
            rows.append({'url':url,'ok':False,'error':str(exc)})
        await asyncio.sleep(0.8)
    return {'ok':True,'results':rows}



@app.get('/api/jobs/{jid}/report')
async def job_report(jid: str, _: bool = Depends(auth)):
    job=await get_job(jid)
    if not job: raise HTTPException(404,'job not found')
    leads=await sb_select('leads',params={'job_id':f'eq.{jid}','select':'*','order':'title_score.desc','limit':'5000'})
    companies=await sb_select('companies',params={'job_id':f'eq.{jid}','select':'root_domain,domain,name,relationship,parent_domain,status,people_found,qualified_people,enriched_people','order':'root_domain.asc,domain.asc','limit':'5000'})
    rels=await sb_select('relationships',params={'job_id':f'eq.{jid}','select':'source_domain,related_domain,related_name,relationship,confidence','order':'confidence.desc','limit':'5000'})
    lines=[f"# Prospecting Beast Report — {jid[:8]}", '', f"Status: {job.get('status','')}", f"Mode: {job.get('mode','')}", f"Leads: {len(leads)}", '', '## Prospects']
    for l in leads:
        rel=l.get('relationship_display') or l.get('relationship') or 'Original company'
        lines.append(f"- **{l.get('name','')}** — {l.get('title','')} → **{l.get('matched_title','')}**; score {l.get('title_score',0)}; company {l.get('company','')}; lineage {rel}; LinkedIn {l.get('linkedin','')}")
    lines += ['', '## Companies']
    for c in companies:
        lines.append(f"- {c.get('domain')} — {c.get('relationship')} — {c.get('status')} — {c.get('qualified_people',0)}/{c.get('people_found',0)} qualified")
    lines += ['', '## Relationships']
    for r in rels:
        lines.append(f"- {r.get('source_domain')} → {r.get('related_domain')} ({r.get('relationship')}, confidence {r.get('confidence',0)})")
    return {'ok':True,'markdown':'\n'.join(lines)}

async def export_rows(jid):
    leads=await sb_select('leads',params={'job_id':f'eq.{jid}','select':'company,domain,relationship,name,title,matched_title,title_score,similarity,linkedin,phone,email,source,seamless_key_index','order':'title_score.desc','limit':'5000'})
    companies=await sb_select('companies',params={'job_id':f'eq.{jid}','select':'root_domain,domain,name,relationship,parent_domain,status,people_found,qualified_people,enriched_people','order':'root_domain.asc,domain.asc','limit':'5000'})
    rels=await sb_select('relationships',params={'job_id':f'eq.{jid}','select':'source_domain,related_domain,related_name,relationship,confidence','order':'confidence.desc','limit':'5000'})
    wb=Workbook(); ws=wb.active; ws.title='Leads'
    lh=['Company','Domain','Root Domain','Parent Domain','Relationship','Relationship Display','Name','Title','Matched Title','Score','Similarity','LinkedIn','Phone','Email','Source','Seamless Key']
    ws.append(lh)
    for r in leads: ws.append([r.get(x,'') for x in ['company','domain','root_domain','parent_domain','relationship','relationship_display','name','title','matched_title','title_score','similarity','linkedin','phone','email','source','seamless_key_index']])
    ws2=wb.create_sheet('Companies'); ws2.append(['Root','Domain','Name','Relationship','Parent Domain','Status','People Found','Qualified','Enriched'])
    for r in companies: ws2.append([r.get(x,'') for x in ['root_domain','domain','name','relationship','parent_domain','status','people_found','qualified_people','enriched_people']])
    ws3=wb.create_sheet('Relationships'); ws3.append(['Source Domain','Related Domain','Related Name','Relationship','Confidence'])
    for r in rels: ws3.append([r.get(x,'') for x in ['source_domain','related_domain','related_name','relationship','confidence']])
    for wsx in wb.worksheets:
        wsx.freeze_panes='A2'; wsx.auto_filter.ref=wsx.dimensions
        for col in wsx.columns:
            width=min(45,max(12,max(len(str(c.value or '')) for c in col)+2))
            wsx.column_dimensions[col[0].column_letter].width=width
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf


@app.get('/api/jobs/{jid}/export')
async def export_job_file(jid: str, _: bool = Depends(auth)):
    job=await get_job(jid)
    if not job: raise HTTPException(404,'job not found')
    buf=await export_rows(jid)
    return StreamingResponse(buf,media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',headers={'Content-Disposition':f'attachment; filename="prospecting_{jid[:8]}.xlsx"'})
